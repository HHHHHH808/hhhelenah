from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

FONT = "Arial"
header_font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", start_color="2F5496")
normal_font = Font(name=FONT, size=11)
wrap = Alignment(vertical="top", wrap_text=True)
header_align = Alignment(vertical="center", horizontal="center", wrap_text=True)
thin = Side(border_style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, ncols, row=1):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

def style_data_row(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font
        cell.alignment = wrap
        cell.border = border

# ---------------- WORK SHEET ----------------
ws = wb.active
ws.title = "Work"

work_headers = [
    "id", "titel", "datum (JJJJ-MM)", "categorieen (komma-gescheiden)", "afbeelding (bestandsnaam/pad)",
    "type (audio/embed/video)", "embedUrl", "videoBestand", "afbeeldingPositie",
    "beschrijving", "nummers (titel - duur, komma-gescheiden, enkel bij type=audio)"
]
ws.append(work_headers)
style_header(ws, len(work_headers))

work_rows = [
    ["wayout", "Andrea Murtas", "", "mixing, production", "assets/Images/Work/wide-circ/blur/circ-wayout.png",
     "audio", "", "", "",
     "Mixing for Andrea Murtas: The Way Out. Client: Sonhouse.",
     "The Way Out - 6:44"],
    ["ww", "Warmste Week 2019", "", "production", "assets/Images/Work/wide-circ/blur/circ-ww.png",
     "audio", "", "", "",
     "Soundtrack for StuBru's promotional campaign for 'De Warmste Week 2019'.",
     "(11 nummers - zie main.js, niet aangepast in dit overzicht)"],
    ["siebert", "Truck", "", "production", "assets/Images/Work/wide-circ/blur/circ-truck.png",
     "embed", "https://player.vimeo.com/video/348057330", "", "",
     "Soundtrack for Siebert Mispelons videowork, 'Truck'.", ""],
    ["burnout", "Burnout", "", "mastering", "assets/Images/Work/wide-circ/blur/circ-burnout.png",
     "embed", "https://www.youtube.com/embed/eq25I4exauE", "", "",
     "Mastering for Epong's soundtrack of Rik Chaubet's film, 'Burnout'.", ""],
    ["ivy", "Ivy Falls", "", "production", "assets/Images/Work/wide-circ/blur/circ-ivy.png",
     "embed", "https://open.spotify.com/embed/track/1iQnPzwIEkNyG0B5zPwIUn", "", "",
     "Production of 'Beautiful Stranger' by Ivy Falls with Kasper Cornelus.", ""],
    ["schob", "Schobbee", "", "mastering", "assets/Images/Work/wide-circ/blur/schob.png",
     "embed", "https://bandcamp.com/EmbeddedPlayer/album=202891063/size=large/bgcol=ffffff/linkcol=de270f/artwork=none/transparent=true/", "", "",
     "Mastering for 'Poe's Law' by Schobbee.", ""],
    ["people", "People's People", "", "production", "assets/Images/Work/wide-circ/blur/circ-people.jpg",
     "embed", "https://www.youtube.com/embed/FD35-NHz7BI", "", "",
     "Sound design and extra arrangements for Lip Service's EP 'People's People'.", ""],
    ["green", "Green Greener Greenest", "", "mastering, mixing, production", "assets/Images/Work/wide-circ/blur/circ-green.jpg",
     "embed", "https://player.vimeo.com/video/221453555", "", "",
     "Sound-design voor Senne Marquenie's short movie 'Green, Greener, Greenest'.", ""],
]

for r in work_rows:
    ws.append(r)

# extra empty rows for new entries
for _ in range(10):
    ws.append([""] * len(work_headers))

for row in range(2, ws.max_row + 1):
    style_data_row(ws, row, len(work_headers))
    ws.cell(row=row, column=3).number_format = "@"

col_widths_work = [12, 24, 14, 22, 45, 16, 45, 18, 18, 50, 40]
for i, w in enumerate(col_widths_work, start=1):
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

ws.freeze_panes = "A2"

# ---------------- LIFE SHEET ----------------
ws2 = wb.create_sheet("Life")
life_headers = [
    "id", "titel", "datum (JJJJ-MM)", "status", "afbeelding (bestandsnaam/pad)",
    "type (audio/embed/video)", "embedUrl", "videoBestand", "afbeeldingPositie",
    "beschrijving"
]
ws2.append(life_headers)
style_header(ws2, len(life_headers))

life_rows = [
    ["comp", "Compressed '16-'17", "", "helenah", "assets/Images/Life/logocompressed.jpg",
     "embed", "https://bandcamp.com/EmbeddedPlayer/album=285491679/size=large/bgcol=ffffff/linkcol=de270f/artwork=none/transparent=true/", "", "",
     "A series of songs made between '16 '17, sonically bundled by various tape & compression emulations."],
    ["hhh", "Helenah", "", "helenah", "assets/Images/Life/logo12.jpg",
     "embed", "https://bandcamp.com/EmbeddedPlayer/album=1065259858/size=large/bgcol=ffffff/linkcol=de270f/artwork=none/transparent=true/", "", "",
     "Debut album on the Brussels based label 'Montage'."],
]
for r in life_rows:
    ws2.append(r)

for _ in range(10):
    ws2.append([""] * len(life_headers))

for row in range(2, ws2.max_row + 1):
    style_data_row(ws2, row, len(life_headers))
    ws2.cell(row=row, column=3).number_format = "@"

col_widths_life = [12, 24, 14, 14, 40, 16, 50, 18, 18, 55]
for i, w in enumerate(col_widths_life, start=1):
    ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = w

ws2.freeze_panes = "A2"

# ---------------- INSTRUCTIONS SHEET ----------------
ws3 = wb.create_sheet("Toelichting", 0)
ws3.sheet_view.showGridLines = False

title_font = Font(name=FONT, bold=True, size=14)
sub_font = Font(name=FONT, bold=True, size=12)
body_font = Font(name=FONT, size=11)

ws3["A1"] = "Toelichting bij dit bestand"
ws3["A1"].font = title_font

lines = [
    ("", ""),
    ("Doel", "sub"),
    ("Vul nieuwe entry's aan onderaan de tabbladen 'Work' en 'Life'. Laat lege rijen leeg "
     "of vul ze aan met je nieuwe projecten. Stuur het bestand terug en de website wordt aangepast.", "body"),
    ("", ""),
    ("Handleiding - stap voor stap", "sub"),
    ("1. Kies het juiste tabblad", "label"),
    ("Ga naar 'Work' voor professionele projecten (mixing/mastering/production/recording), "
     "of naar 'Life' voor eigen muziek/persoonlijke releases.", "body"),
    ("2. Vul een nieuwe rij in (of vul een lege rij aan)", "label"),
    ("Onderaan elk tabblad staan lege rijen klaar. Vul per kolom de gevraagde info in. "
     "Zie hieronder bij 'Kolommen uitgelegd' wat elke kolom betekent en welke waarden geldig zijn. "
     "Niet elke kolom is verplicht - laat leeg wat niet van toepassing is.", "body"),
    ("3. Bezorg de bijhorende bestanden", "label"),
    ("Als je nieuwe afbeeldingen, audio (mp3) of video toevoegt, stuur deze bestanden gewoon mee "
     "(bv. los, of in een mapje). Geef in de Excel-kolom enkel de bestandsnaam op (bv. 'circ-nieuwproject.jpg'), "
     "ik plaats het bestand zelf op de juiste plek op de website.", "body"),
    ("4. Stuur het bestand terug", "label"),
    ("Stuur dit volledige Excel-bestand terug. Ik verwerk de nieuwe/aangepaste rijen dan in de website.", "body"),
    ("5. Controle achteraf", "label"),
    ("Eens verwerkt, laat ik je een voorbeeld zien zodat je kan controleren of alles er goed uitziet "
     "voor je het op de live site zet.", "body"),
    ("", ""),
    ("Kolommen uitgelegd", "sub"),
    ("id", "label"),
    ("Korte unieke code, alleen kleine letters/cijfers, geen spaties (bv. 'nieuwproject'). "
     "Wordt intern gebruikt door de website, niet zichtbaar voor bezoekers.", "body"),
    ("titel", "label"),
    ("De titel/naam die zichtbaar is op de website.", "body"),
    ("datum", "label"),
    ("Formaat JJJJ-MM (bv. 2024-03 voor maart 2024). Wordt gebruikt om de items chronologisch "
     "te ordenen. Als je enkel het jaar weet, mag je JJJJ-01 gebruiken. Laat leeg als onbekend. "
     "Let op: deze kolom staat als tekst geformatteerd, zodat Excel er geen datum van maakt.", "body"),
    ("categorieen", "label"),
    ("Eén of meerdere van: production, mixing, mastering, recording. Komma-gescheiden bij meerdere.", "body"),
    ("status (enkel Life)", "label"),
    ("Groepering voor het 'Life'-tabblad, bv. 'helenah'.", "body"),
    ("afbeelding", "label"),
    ("Bestandsnaam (en pad) van de afbeelding. Plaats de afbeelding zelf in de juiste map "
     "(assets/Images/Work/... of assets/Images/Life/...) en geef hier het pad op. "
     "Werk-afbeeldingen zijn best vierkant (worden in een cirkel getoond).", "body"),
    ("type", "label"),
    ("'audio' = lijst van eigen nummers (mp3's op de site zelf), "
     "'embed' = ingesloten speler (Bandcamp/Spotify/Vimeo), "
     "'video' = eigen videobestand.", "body"),
    ("embedUrl", "label"),
    ("Enkel invullen bij type='embed'. Plak hier de volledige embed-URL:\n"
     " - Bandcamp: gebruik de 'EmbeddedPlayer' URL met 'size=large'\n"
     " - Spotify: open.spotify.com/embed/...\n"
     " - Vimeo: player.vimeo.com/video/...\n"
     " - YouTube: LET OP - YouTube embeds geven vaak 'error 153' als de eigenaar embedden "
     "heeft uitgeschakeld. Gebruik bij voorkeur Vimeo of Bandcamp, of laat dit veld leeg "
     "als je geen embed hebt.", "body"),
    ("videoBestand", "label"),
    ("Enkel invullen bij type='video'. Bestandsnaam van het videobestand.", "body"),
    ("afbeeldingPositie", "label"),
    ("Optioneel - meestal mag je dit gewoon LEEG laten. Op de website wordt elke afbeelding "
     "bijgesneden tot een vast formaat (vierkant bij Life, cirkel bij Work), waarbij standaard "
     "het midden van de foto getoond wordt. Als bij een nieuwe foto net het belangrijkste deel "
     "(bv. iemands gezicht) wegvalt door dat bijsnijden, kan je hier aangeven welk deel zichtbaar "
     "moet blijven. Geldige waarden zijn bv.: 'center top' (toon bovenkant), 'center bottom' "
     "(toon onderkant), 'left center', 'right center', of een percentage zoals '50% 20%'. "
     "Twijfel je? Laat leeg - ik pas dit later eventueel aan als de foto er verkeerd uitgesneden uitziet.", "body"),
    ("beschrijving", "label"),
    ("De tekst die getoond wordt bij het item.", "body"),
    ("nummers (enkel Work, type=audio)", "label"),
    ("Lijst van eigen nummers in formaat 'Titel - duur', komma-gescheiden "
     "(bv. 'Intro - 2:30, Outro - 3:10'). De mp3-bestanden zelf moeten apart aangeleverd worden.", "body"),
    ("", ""),
    ("Belangrijk", "sub"),
    ("- Voeg nieuwe rijen toe onderaan de tabel, of vul de lege rijen in.\n"
     "- Verwijder geen kolommen.\n"
     "- Voeg afbeeldingen en eventuele audio/video bestanden apart toe (los van dit Excel-bestand) "
     "en vermeld de bestandsnaam in de juiste kolom.", "body"),
]

row_idx = 3
for text, kind in lines:
    cell = ws3.cell(row=row_idx, column=1, value=text)
    if kind == "sub":
        cell.font = sub_font
    elif kind == "label":
        cell.font = Font(name=FONT, bold=True, size=11)
    else:
        cell.font = body_font
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    if text:
        ws3.row_dimensions[row_idx].height = max(15, 15 * (text.count("\n") + 1) * (1 + len(text)//90))
    row_idx += 1

ws3.column_dimensions["A"].width = 110

wb.save("/Users/martijndebondt/Documents/Github/hhhelenah/exports/hhhelenah_entries.xlsx")
print("done")
